"""
Accounting Routes - Admin Financial Dashboard
Complete financial tracking and reporting API
"""

from datetime import datetime, timedelta, timezone
from decimal import Decimal
from typing import List, Optional

from api.database.connection import get_db
from api.middleware.auth import get_current_user, require_admin
from api.services.accounting import AccountingService
from fastapi import APIRouter, Depends, HTTPException, Query, status
from pydantic import BaseModel, Field

router = APIRouter()


# =============================================================================
# RESPONSE MODELS
# =============================================================================

class DashboardMetrics(BaseModel):
    """Main dashboard metrics"""
    period_days: int
    total_revenue: float
    total_costs: float
    openai_costs: float
    gross_profit: float
    profit_margin_percent: float
    credits_sold: int
    credits_consumed: int
    transactions_count: int
    avg_transaction_value: float


class DailyRevenueItem(BaseModel):
    """Daily revenue data point"""
    date: str
    revenue: float
    costs: float
    profit: float
    transactions: int


class OperationsBreakdown(BaseModel):
    """Operations by type"""
    copy_generation: int = 0
    trend_analysis: int = 0
    niche_report: int = 0
    ai_chat: int = 0
    image_generation: int = 0


class TopUser(BaseModel):
    """Top spending user"""
    user_id: str
    total_spent: float
    credits_purchased: int
    credits_used: int
    purchase_count: int
    lifetime_profit: float


class PackageSales(BaseModel):
    """Package sales stats"""
    name: str
    slug: str
    sales: int
    revenue: float
    credits: int


class CreditPackageResponse(BaseModel):
    """Credit package for display"""
    id: str
    name: str
    slug: str
    credits: int
    price_brl: float
    price_per_credit: float
    original_price: Optional[float]
    discount_percent: int
    description: Optional[str]
    badge: Optional[str]
    is_featured: bool


class FinancialSummary(BaseModel):
    """Complete financial summary"""
    metrics: DashboardMetrics
    revenue_by_day: List[DailyRevenueItem]
    operations_breakdown: OperationsBreakdown
    top_users: List[TopUser]
    package_sales: List[PackageSales]


# =============================================================================
# DASHBOARD ENDPOINTS
# =============================================================================

@router.get("/dashboard", response_model=DashboardMetrics)
async def get_dashboard_metrics(
    days: int = Query(30, ge=1, le=365, description="Period in days"),
    current_user = Depends(require_admin),
    db = Depends(get_db)
):
    """
    Get main dashboard metrics for admin panel.
    Requires admin authentication.
    """
    service = AccountingService(db)
    metrics = await service.get_dashboard_metrics(days)
    return DashboardMetrics(**metrics)


@router.get("/revenue/daily", response_model=List[DailyRevenueItem])
async def get_daily_revenue(
    days: int = Query(30, ge=1, le=365),
    current_user = Depends(require_admin),
    db = Depends(get_db)
):
    """Get daily revenue breakdown for charts"""
    service = AccountingService(db)
    data = await service.get_revenue_by_day(days)
    return [DailyRevenueItem(**item) for item in data]


@router.get("/operations/breakdown", response_model=OperationsBreakdown)
async def get_operations_breakdown(
    days: int = Query(30, ge=1, le=365),
    current_user = Depends(require_admin),
    db = Depends(get_db)
):
    """Get breakdown of operations by type"""
    service = AccountingService(db)
    data = await service.get_operations_breakdown(days)
    return OperationsBreakdown(**data)


@router.get("/users/top", response_model=List[TopUser])
async def get_top_users(
    limit: int = Query(10, ge=1, le=100),
    current_user = Depends(require_admin),
    db = Depends(get_db)
):
    """Get top spending users"""
    service = AccountingService(db)
    users = await service.get_top_users(limit)
    return [TopUser(**u) for u in users]


@router.get("/packages/sales", response_model=List[PackageSales])
async def get_package_sales(
    days: int = Query(30, ge=1, le=365),
    current_user = Depends(require_admin),
    db = Depends(get_db)
):
    """Get sales stats per package"""
    service = AccountingService(db)
    sales = await service.get_package_sales_stats(days)
    return [PackageSales(**s) for s in sales]


@router.get("/summary", response_model=FinancialSummary)
async def get_financial_summary(
    days: int = Query(30, ge=1, le=365),
    current_user = Depends(require_admin),
    db = Depends(get_db)
):
    """
    Get complete financial summary for dashboard.
    Combines all metrics in a single request.
    """
    service = AccountingService(db)
    
    metrics = await service.get_dashboard_metrics(days)
    revenue_by_day = await service.get_revenue_by_day(days)
    operations = await service.get_operations_breakdown(days)
    top_users = await service.get_top_users(10)
    package_sales = await service.get_package_sales_stats(days)
    
    return FinancialSummary(
        metrics=DashboardMetrics(**metrics),
        revenue_by_day=[DailyRevenueItem(**item) for item in revenue_by_day],
        operations_breakdown=OperationsBreakdown(**operations),
        top_users=[TopUser(**u) for u in top_users],
        package_sales=[PackageSales(**s) for s in package_sales],
    )


# =============================================================================
# CREDIT PACKAGES ENDPOINTS
# =============================================================================

@router.get("/packages", response_model=List[CreditPackageResponse])
async def list_credit_packages(
    db = Depends(get_db)
):
    """
    List all active credit packages.
    Public endpoint for checkout page.
    """
    service = AccountingService(db)
    packages = await service.get_active_packages()
    
    return [
        CreditPackageResponse(
            id=str(pkg.id),
            name=pkg.name,
            slug=pkg.slug,
            credits=pkg.credits,
            price_brl=float(pkg.price_brl),
            price_per_credit=float(pkg.price_per_credit),
            original_price=float(pkg.original_price) if pkg.original_price else None,
            discount_percent=pkg.discount_percent or 0,
            description=pkg.description,
            badge=pkg.badge,
            is_featured=pkg.is_featured,
        )
        for pkg in packages
    ]


# =============================================================================
# REPORTS ENDPOINTS
# =============================================================================

@router.post("/reports/generate-daily")
async def generate_daily_report(
    date: Optional[datetime] = None,
    current_user = Depends(require_admin),
    db = Depends(get_db)
):
    """Manually trigger daily report generation"""
    service = AccountingService(db)
    report = await service.generate_daily_report(date)
    return {
        "success": True,
        "report_date": str(report.report_date),
        "total_revenue": float(report.total_revenue),
        "gross_profit": float(report.gross_profit),
    }


@router.get("/reports/export")
async def export_financial_report(
    start_date: datetime = Query(...),
    end_date: datetime = Query(...),
    format: str = Query("csv", pattern="^(csv|xlsx|json)$"),
    current_user = Depends(require_admin),
    db = Depends(get_db)
):
    """Export financial report in various formats"""
    import csv
    import json
    from io import BytesIO, StringIO

    from api.database.accounting_models import DailyReport
    from fastapi.responses import StreamingResponse
    # Buscar dados do relatório
    from sqlalchemy import func, select
    
    result = await db.execute(
        select(DailyReport).where(
            DailyReport.report_date >= start_date.date(),
            DailyReport.report_date <= end_date.date()
        ).order_by(DailyReport.report_date)
    )
    reports = result.scalars().all()
    
    if format == "json":
        # Export como JSON
        data = [
            {
                "date": str(r.report_date),
                "total_revenue": float(r.total_revenue),
                "subscription_revenue": float(r.subscription_revenue),
                "credits_revenue": float(r.credits_revenue),
                "gross_profit": float(r.gross_profit),
                "openai_cost": float(r.openai_cost),
                "active_users": r.active_users,
                "new_subscriptions": r.new_subscriptions,
            }
            for r in reports
        ]
        
        output = BytesIO()
        output.write(json.dumps(data, indent=2).encode('utf-8'))
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename=report_{start_date.date()}_{end_date.date()}.json"
            }
        )
    
    elif format == "csv":
        # Export como CSV
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "Data", "Receita Total", "Assinaturas", "Créditos",
            "Lucro Bruto", "Custo OpenAI", "Usuários Ativos", "Novas Assinaturas"
        ])
        
        # Dados
        for r in reports:
            writer.writerow([
                str(r.report_date),
                float(r.total_revenue),
                float(r.subscription_revenue),
                float(r.credits_revenue),
                float(r.gross_profit),
                float(r.openai_cost),
                r.active_users,
                r.new_subscriptions,
            ])
        
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=report_{start_date.date()}_{end_date.date()}.csv"
            }
        )
    
    elif format == "xlsx":
        # Export como Excel (requer openpyxl)
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório Financeiro"
            
            # Header
            headers = [
                "Data", "Receita Total", "Assinaturas", "Créditos",
                "Lucro Bruto", "Custo OpenAI", "Usuários Ativos", "Novas Assinaturas"
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Dados
            for row, r in enumerate(reports, 2):
                ws.cell(row=row, column=1, value=str(r.report_date))
                ws.cell(row=row, column=2, value=float(r.total_revenue))
                ws.cell(row=row, column=3, value=float(r.subscription_revenue))
                ws.cell(row=row, column=4, value=float(r.credits_revenue))
                ws.cell(row=row, column=5, value=float(r.gross_profit))
                ws.cell(row=row, column=6, value=float(r.openai_cost))
                ws.cell(row=row, column=7, value=r.active_users)
                ws.cell(row=row, column=8, value=r.new_subscriptions)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=report_{start_date.date()}_{end_date.date()}.xlsx"
                }
            )
            
        except ImportError:
            raise HTTPException(
                status_code=501,
                detail="Excel export requires openpyxl. Use CSV or JSON."
            )


# =============================================================================
# COST TRACKING ENDPOINTS
# =============================================================================

@router.get("/costs/openai")
async def get_openai_costs(
    days: int = Query(30, ge=1, le=365),
    current_user = Depends(require_admin),
    db = Depends(get_db)
):
    """Get OpenAI API costs breakdown"""
    # Query API usage logs
    from api.database.accounting_models import APIUsageLog
    from sqlalchemy import func, select
    
    start_date = datetime.now(timezone.utc) - timedelta(days=days)
    
    result = await db.execute(
        select(
            func.sum(APIUsageLog.cost_brl).label("total_cost"),
            func.sum(APIUsageLog.tokens_input).label("total_input"),
            func.sum(APIUsageLog.tokens_output).label("total_output"),
            func.count(APIUsageLog.id).label("request_count"),
            func.avg(APIUsageLog.request_duration_ms).label("avg_duration"),
        ).where(APIUsageLog.created_at >= start_date)
    )
    row = result.fetchone()
    
    return {
        "period_days": days,
        "total_cost_brl": float(row.total_cost or 0),
        "total_tokens_input": int(row.total_input or 0),
        "total_tokens_output": int(row.total_output or 0),
        "request_count": int(row.request_count or 0),
        "avg_request_duration_ms": float(row.avg_duration or 0),
    }


@router.get("/costs/by-operation")
async def get_costs_by_operation(
    days: int = Query(30, ge=1, le=365),
    current_user = Depends(require_admin),
    db = Depends(get_db)
):
    """Get costs breakdown by operation type"""
    from api.database.accounting_models import APIUsageLog
    from sqlalchemy import func, select
    
    start_date = datetime.now(timezone.utc) - timedelta(days=days)
    
    result = await db.execute(
        select(
            APIUsageLog.operation_type,
            func.sum(APIUsageLog.cost_brl).label("cost"),
            func.count(APIUsageLog.id).label("count"),
            func.sum(APIUsageLog.tokens_total).label("tokens"),
        )
        .where(APIUsageLog.created_at >= start_date)
        .group_by(APIUsageLog.operation_type)
    )
    
    return [
        {
            "operation": row.operation_type,
            "cost_brl": float(row.cost or 0),
            "request_count": row.count,
            "total_tokens": row.tokens,
        }
        for row in result.fetchall()
    ]
